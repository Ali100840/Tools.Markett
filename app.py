from flask import Flask, render_template, request, jsonify
from datetime import datetime
import openpyxl
from openpyxl import Workbook

app = Flask(__name__)

products = {
    0: ["saw", 10],
    1: ["drill", 30],
    2: ["hammer", 15]
}

@app.route("/")
def index():
    return render_template("index.html", products=products)

from openpyxl import load_workbook

@app.route("/generate-bill", methods=["POST"])
def generate_bill():
    data = request.json
    quantities = data["quantities"]
    name = data.get("name", "")
    phone = data.get("phone", "")
    location = data.get("location", "")
    bill = []
    total_all = 0

    for i, qty in enumerate(quantities):
        qty = int(qty)
        if qty > 0:
            product_name, price = products[i]
            total = qty * price
            bill.append({
                "name": product_name,
                "price": price,
                "qty": qty,
                "total": total
            })
            total_all += total

    # فتح ملف إكسل موجود
 excel_file = '/tmp/ALI(2).xlsx'  # تأكد من استخدام /tmp

# تحقق إذا الملف موجود
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    # أضف رؤوس الأعمدة فقط إذا الملف جديد
    ws.append(["Name", "Phone", "Location", "Total", "Qty", "Price", "Product", "Date"])

# هذا الكود لازم يتنفذ دائمًا، مش فقط لو الملف جديد
for item in bill:
    ws.append([
        name,
        phone,
        location,
        item["total"],
        item["qty"],
        item["price"],
        item["name"],
        datetime.now().strftime("%Y-%m-%d")
    ])

# احفظ الملف دايمًا
wb.save(excel_file)

    # حفظ الملف بعد التعديل
    wb.save('ALI(2).xlsx')

    return jsonify({
        "bill": bill,
        "total": total_all,
        "date": datetime.now().strftime("%d-%m-%Y"),
        "name": name,
        "phone": phone,
        "location": location
    })

            
            






def save():

    pass

if __name__ == "__main__":
    app.run(app.run(host="0.0.0.0", port=3000)
    )
